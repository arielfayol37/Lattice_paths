# python3 run.py
import os
import shelve
import pickle
import openpyxl, time, concurrent.futures as cf
from Population import Population
from Sequence import Sequence
from lp_utils import generate_all_paths
from pebble import ProcessPool
from Genome import Genome



def collect_data_genetic(m, n):
    """
    Runs the genetic algorithm on lattice using the values generated by the greedy algorithm as initial target values.
    Then incrementing those values until the program does not find a solution.
    """
    wb1 = openpyxl.load_workbook(
        "lattice_table_greedy_double_slicing_rloo" + str(n) + ".xlsx"
    )  # Open greedy algorithm spreadsheets
    wb2 = openpyxl.load_workbook(
        "lattice_table_greedy_double_slicing_loo" + str(n) + ".xlsx"
    )
    filename = "lattice_table_genetic_" + str(n) + ".xlsx"
    config_indices = (
        []
    )  # To store the configuration indices of the genetic algorithm that returns the solution the quickest.
    shelfFile = shelve.open("evolution_config")
    try:
        wb = openpyxl.load_workbook(
            filename
        )  # Open the spreadsheet for the genetic algorithm.
    except:
        wb = openpyxl.Workbook()

    for i in range(
        m, 12
    ):  # we stop at m == 11 just like the tables in the previous papers
        for k in range(3, i + n - 1):
            if (
                wb1["Sheet"].cell(i + 1, k + 1).value
                > wb2["Sheet"].cell(i + 1, k + 1).value
            ):  # Checking which of the version of the greedy algorithms have
                # a greater value at this cell
                target = wb1["Sheet"].cell(i + 1, k + 1).value
            else:
                target = wb2["Sheet"].cell(i + 1, k + 1).value
            print(f"Starting parallel search for target = {target} m = {i}, n = {n}, and k = {k}...")
            run, ci = parallel_search(target=target, m=i, n=n, k=k)
            print(f"Done parallel search for target = {target} m = {i}, n = {n}, and k = {k}")
            try:  # Might raise an error if the cell does not have a value yet
                if int(wb["Sheet"].cell(i + 1, k + 1).value) < target:
                    wb["Sheet"].cell(i + 1, k + 1).value = target
            except:
                wb["Sheet"].cell(i + 1, k + 1).value = target
            wb.save(filename)
            while (
                run
            ):  # Keeps incrementing the target and stops only if it didn't find any solution.
                config_indices.append(ci)
                target += 1
                print(f"Starting parallel search for target = {target} m = {i}, n = {n}, and k = {k}...")
                run, ci = parallel_search(target=target, m=i, n=n, k=k)
                print(f"Done parallel search for target = {target} m = {i}, n = {n}, and k = {k}")
            try:
                if int(wb["Sheet"].cell(i + 1, k + 1).value) < target-1:# must be minus one here because run returned False
                    # in the while loop above, meaning solution not found
                    wb["Sheet"].cell(i + 1, k + 1).value = target-1
            except:
                wb["Sheet"].cell(i + 1, k + 1).value = target-1 # in case excel cell is empty
            wb.save(filename)
            shelfFile[
                "list_of_configs"
            ] = config_indices  # Store the config indices in a shelf file
            for config in config_indices:
                save_config_in_txt_file(config,target,m,n,k)
    
    wb.close()
    wb1.close()
    wb2.close()


def search(
    size,
    target,
    m,
    n,
    k,
    kill_mode="non_bias_random",
    mode="roulette",
    norm=True,
    scale=True,
    draw=False,
    visualize=False,
    temp=4.0,
):
    """
    Initializes a population and searchs for the perfect individual with j genes given m, n, and k.
    m: number of rows of lattice
    n: number of columns of lattice
    k: if two paths share k number of edges or more, then they are said to be k-equivalent.
    temp: how fast the population is allowed to converge
    visualize: if True, plot the change the in average population fitness, divergence, and maximum fitness.

    """

    world = Population(size, target, m, n, k, norm=norm, scale=scale, temp=temp)
    best = world.evolve(mode, kill_mode)
    if draw:
        best.draw()
    if visualize:
        world.visualize_evolution()
    return world

def previous_population_saved(target, m, n, k):
    """
    Used to avoid overwrites in parallel_search()
    """
    try:
        filename = "lattice_table_genetic_" + str(n) + ".xlsx"
        wb = openpyxl.load_workbook(
            filename
        ) 
        if int(wb["Sheet"].cell(m + 1, k + 1).value) >= target:# must be minus one here
            return True # Do not save
        else:
            return False # Can save
    except:
        return False # Empty excel cell, meaning nothing was saved before    

def parallel_search(target, m, n, k):
    """
    Use different populations with different popularion sizes and temperature values.
    Run search() with these configurations in parallel. Whenever one of the populations return
    the solution, stop and return True and the index of the configuration that found the solution.
    If none of the parallel searches find the solution, then we combine the poulations, and run the
    search on the combine population. If the solution is still not found, we return False.
    """
    # Set the parameter values for the search function
    evolution_parameters = [
        {
            "size": 1000,
            "target": target,
            "m": m,
            "n": n,
            "k": k,
            "kill_mode": "non_bias_random",
            "temp": 5,
        },
        {
            "size": 2000,
            "target": target,
            "m": m,
            "n": n,
            "k": k,
            "kill_mode": "non_bias_random",
            "temp": 10,
        },
        {
            "size": 5000,
            "target": target,
            "m": m,
            "n": n,
            "k": k,
            "kill_mode": "non_bias_random",
            "temp": 7,
        },
        {
            "size": 10000,
            "target": target,
            "m": m,
            "n": n,
            "k": k,
            "kill_mode": "non_bias_random",
            "temp": 4,
        },
    ]

    shelfFile = shelve.open("populations_genetic_data")
    config_index = (
        -1
    )  # Configuration index to be returned if none of the parallel searches finds the solution
    recheck = True # To solve a the problem in if taks[-1].done()
    with ProcessPool() as executor:
        # Create a list of tasks
        tasks = []
        for param in evolution_parameters:
            task = executor.schedule(search, kwargs=param)
            tasks.append(task)

        population = (
            "population_" + str(target) + "_" + str(m) + "_" + str(n) + "_" + str(k)
        )
        print("Entering while loop")
        tasks_done = []
        # Wait for the tasks to complete
        run = True
        while run:
            for i in range(len(tasks)):
                if tasks[i].done():
                    if i not in tasks_done:
                        tasks_done.append(i)
                        print(str(i), "th task done")
                        result = tasks[i].result()
                        
                        if result.fitnesses[result.bfi] == 9999:
                            print("Perfect individual found")
                            shelfFile[population] = result
                            save_object(result, population)
                            run = False
                            config_index = i
                            for j in range(len(tasks)):
                                c = tasks[
                                    j
                                ].cancel()  # will return False if already completed, and will cancel and return True otherwise
                            break  # breaking out of for loop
            if run != False:
                #print("sleeping...")
                time.sleep(target * 10)

            if tasks[-1].done():
                
                # task[-1].done() may return True, while result has not yet been assigned. This may happen if the task
                # gets completed while the program was sleeping.
                # Problem solved using recheck
                 
                if recheck:
                    recheck = False
                    continue # This is actually useless, but just in case we do further modifications and forget that we 
                # have to recheck and not assign False to run
                
                else:
                    
                    try:
                        print(
                            type(result)
                        )  # checking whether result has already been assigned. This will raise an error if result has not yet been
                    # assigned
                    except:
                        print("Last task done, no perfect individual found. Will merge populations")
                        print("getting the first population")
                        result = tasks[
                            0
                        ].result()  # if result was not assigned, then none of the tasks found the solution.
                    # Don't put 'pass' here because it will create an error(bug) on the next line
                    if (
                        result.fitnesses[result.bfi] != 9999
                    ):  # important because even if task[-1].done() returns True, it might return true just because
                        # it was cancelled.

                        print("Merging populations")
                        new_pop = tasks[0].result()  # the search() returns a Population
                        for task in tasks[1:]:
                            new_pop.individuals += task.result().individuals
                            new_pop.fitnesses += task.result().fitnesses
                            new_pop.max_size = 2000
                            new_pop.av_pop_fitnesses.clear()
                            new_pop.av_pop_divergences.clear()
                            new_pop.roulette_ready = False
                        print("Merged population evolving")
                        new_best_individual = new_pop.evolve(
                            mode="roulette", kill_mode="non_bias_random"
                        )
                        run = False  # To break out of the while loop
                        result = new_pop
                        if not previous_population_saved(target, m, n, k):
                            shelfFile[population] = result
                            save_object(result, population)

    print("Done with all tasks")
    result.individuals[result.bfi].show(result.paths)
    f = result.fitnesses[result.bfi]
    shelfFile.close()
    print("BestFitness:", f, "/", result.fm, " Config_index: ", config_index)
    if f == 9999:
        return (True, config_index)
    else:
        return (False, config_index)


def greedy(m, n, k, pats, reverse=True, return_solution = False, pop_out_indices = []):
    """
    Greedy algorithm: Takes a set of paths 'pats' for an m by n lattice. Then creates another set G.
    Reverse: if False, it goes through every path in pats and checks wheter the path is k-distinct from
             every path in G. if the case, then add the path to G.
             if True, iterate through the paths in pats in reverse.
    Returns the size of G (the number of paths in G)
    """
    sol = []  # The set G
    solutions_indices = [] # Watch out here because some paths are sometimes pop out before the list
                            # of paths is passed as argument
    if not reverse:
        l = Sequence(m, n, empty=True)
        si = 0 # start index
        while si in pop_out_indices: # implemented this just in case we wanted to quickly run the
                                     # greedy algorithm without some paths
            si += 1
        l.terms = pats[si]
        sol.append(l)
        solutions_indices.append(si)
        for index,pat in enumerate(pats[:]):
            if index in pop_out_indices:
                continue
            s = Sequence(m, n, empty=True)
            s.terms = pat
            
            for i in range(len(sol)):
                if s.compare(sol[i], k) == 0:
                    break
                if i == len(sol) - 1:
                    sol.append(s)
                    solutions_indices.append(index)
    else:
        l = Sequence(m, n, empty=True)
        si = len(pats)-1
        while si in pop_out_indices:
            si -= 1
        l.terms = pats[si]
        sol.append(l)
        solutions_indices.append(si)
        for j in range(len(pats) - 1, -1, -1):
            if j in pop_out_indices:
                continue
            s = Sequence(m, n, empty=True)
            s.terms = pats[j]
            for i in range(len(sol)):
                if s.compare(sol[i], k) == 0:
                    break
                if i == len(sol) - 1:
                    sol.append(s)
                    solutions_indices.append(j) 
    if return_solution:
        return (len(sol), solutions_indices)
    else:
        return len(sol)


def greedy_t(m, n, k, paths, reverse):
    """
    Improved greedy algorithm: Run the greedy algorithm, but after going through all the paths once, iterate again
    through the paths, and for every iteration, run the greedy algorithm without one of the paths. Then we do it again
    without 2 of the paths. So we try every combination of two paths we cannot remove and run the greedy algorithm.
    Returns a list values returned by the greedy algorithm on the different set of paths.
    """
    solutions = []
    pati = (
        paths.copy()
    )  # We use copy here because we don't want pati and paths to have the same reference
    max_indices = (
        -2,
        -2,
    )  # This stores the indices of the paths we removed that returned the greatest result from greedy()
    max = greedy(m, n, k, pati, reverse)  # we run the greedy() first with all the paths
    solutions.append(max)
    for i in range(len(paths)):
        patos = (
            paths.copy()
        )  # paths, not pati because pati might have been been reversed inside greedy()
        patos.pop(i)  # We remove the current path
        for j in range(len(patos)):
            patz = patos.copy()
            patz.pop(j)
            amax = greedy(
                m, n, k, patz, reverse
            )  # Run the greedy with two paths popped out
            if amax > max:
                max = amax
                if j >= i:
                    max_indices = (i, j + 1)
                else:
                    max_indices = (i, j)
            solutions.append(max)

        assert len(patos) == len(paths) - 1
        pmax = greedy(m, n, k, patos, reverse)
        # Very important that the following comes last.
        # Otherwise we would think popping out two paths is necessary,
        # whereas only one was enough
        if pmax > max:
            max = pmax
            max_indices = (i, -1) # if popping only one path was enough
        solutions.append(pmax)
    if max_indices[1] == -2:
        foo, solution_indices = greedy(m,n,k, paths.copy(), reverse, return_solution= True)
    elif max_indices[1] == -1:
        patos = paths.copy()
        patos.pop(max_indices[0])
        foo, solution_indices = greedy(m,n,k,patos,reverse, return_solution=True)
    else:
        patos = paths.copy()
        patos.pop(max_indices[0])
        if (max_indices[1]) < max_indices[0]:
            patos.pop(max_indices[1])
        elif (max_indices[1] - 1) >= max_indices[0]:
            patos.pop(max_indices[1] - 1)
        foo, solution_indices = greedy(m,n,k,patos,reverse, return_solution=True)
    # print(solution_indices)
    g = Genome(num_sequences=len(solution_indices), m = m, n = n, k = k, paths=None, len_paths=len(paths), empty=True)
    g.sequences = solution_indices
    pathsii = list(map(lambda alphabet_path:"".join(alphabet_path), g.translate(paths)))
    filename_start = "simple_" if max_indices[1] < 0 else "improved_"
    filename_start += "reversed" if reverse else "normal_order"   
    filename = f"{filename_start}_greedy_solution_{len(solution_indices)}_{m}_{n}_{k}.txt"
    with open(filename, "a") as file:
        file.write(filename + " equivalent to greedyAlgorithmType_target_m_n_k.txt\n")
        for index, path in enumerate(pathsii):
            file.write(path + " " + str(solution_indices[index]) + '\n')
        file.write("\n\n\n")
    return (solutions, max_indices)


def find_max(alist):
    """
    Finding the maximum value in a list.
    Returns the maximum value in alist.
    """
    maxs = alist[0]
    for item in alist:
        if item > maxs:
            maxs = item
    return maxs


def collect_data_greedy(m, n):
    """
    Runs the improved greedy algorithm, greedy_t(), and store the values in excel spreadsheets.
    One excel spreadsheat will store the data for lexicographical order, and the other will store the data
    for the reverse run.
    No value returned.
    """
    filename1 = (
        "lattice_table_greedy_double_slicing_rloo" + str(n) + ".xlsx"
    )  # spreadsheet name for running the greedy algorithm in reverse lexicographical order
    try:
        wb1 = openpyxl.load_workbook(filename1)
    except:
        wb1 = openpyxl.Workbook()
    filename2 = (
        "lattice_table_greedy_double_slicing_loo" + str(n) + ".xlsx"
    )  # spreadsheet name for running the greedy algorithm in lexicographical order
    try:
        wb2 = openpyxl.load_workbook(filename2)
    except:
        wb2 = openpyxl.Workbook()
    filename3 = (
        "lattice_table_simple_greedy" + str(n) + ".xlsx"
    )  # spreadsheet name for running the greedy algorithm in lexicographical order
    try:
        wb3 = openpyxl.load_workbook(filename3)
    except:
        wb3 = openpyxl.Workbook()    
    sheet1 = wb1["Sheet"]
    sheet1.cell(1, 1).value = "m/k"
    sheet2 = wb2["Sheet"]
    sheet2.cell(1, 1).value = "m/k"
    sheet3 = wb3["Sheet"]
    sheet3.cell(1, 1).value = "m/k"

    for i in range(m, 12):
        paths = generate_all_paths(i, n) # This seems kinda stupid because the paths are generated from the arguments
        # of greedy() itself but the issue is that we sometimes want to pop some paths

        for k in range(3, i + n - 1):
            sheet1.cell(1, k + 1).value = str(k)
            sheet1.cell(i + 1, 1).value = str(i)
            sheet1.cell(1 + 15, 2 * k + 1).value = str(k)
            sheet1.cell(i + 1 + 15, 1).value = str(i)
            f1, max_indexes1 = greedy_t(i, n, k, paths, True)

            maxi1 = find_max(f1)
            sheet1.cell(i + 1, k + 1).value = maxi1
            sheet1.cell(i + 1 + 15, 2 * k + 1).value = max_indexes1[0]
            sheet1.cell(i + 1 + 15, 2 * k + 2).value = max_indexes1[1]

            wb1.save(filename1)

            sheet2.cell(1, k + 1).value = str(k)
            sheet2.cell(i + 1, 1).value = str(i)
            sheet2.cell(1 + 15, 2 * k + 1).value = str(k)
            sheet2.cell(i + 1 + 15, 1).value = str(i)
            f2, max_indexes2 = greedy_t(i, n, k, paths, False)

            maxi2 = find_max(f2)
            sheet2.cell(i + 1, k + 1).value = maxi2
            sheet2.cell(i + 1 + 15, 2 * k + 1).value = max_indexes2[0]
            sheet2.cell(i + 1 + 15, 2 * k + 2).value = max_indexes2[1]


            wb2.save(filename2)

            sheet3.cell(1, k + 1).value = str(k)
            sheet3.cell(i + 1, 1).value = str(i)
            sheet3.cell(i + 1, k + 1).value = greedy(i,n,k,paths,reverse=False)

            wb3.save(filename3)
            

# The next functions were used during the development of the program
# but aren't used anymore. However, I kept them here just in case.

def greedy_recursion(m,n,k,paths,recurse=False):
    solutions = []
    pati = paths.copy()
    solutions.append(greedy(m,n,k,pati))
    for i in range(len(paths)):
        patos = paths.copy()
        patos.pop(i)
        if recurse:
            patos2 = patos.copy()
            
             
            solutions+=greedy_recursion(m,n,k,patos2,False)
           
        assert len(patos) == len(paths)-1
        
        solutions.append(greedy(m,n,k,patos))
    return solutions

def test(
    size,
    target,
    m,
    n,
    k,
    kill_mode="non_bias_random",
    mode="roulette",
    norm=True,
    scale=True,
    draw=False,
    visualize=False,
    temp=4.0,
):
    """
    Test function for evolution process.
    No return value.
    """
    start_time = time.perf_counter()
    world = Population(size, target, m, n, k, norm=norm, scale=scale, temp=temp)

    best = world.evolve(mode, kill_mode)
    end_time = time.perf_counter()
    print(f"It took {(end_time-start_time)/60} minutes to find this solution")
    if draw:
        best.draw()
    if visualize:
        world.visualize_evolution()
    return

def save_object(obj, filename):
    """
    Function to save the populations in a pickle file, since shelve may
    not work on certain computers
    """
    with open(filename, 'wb') as outp:  # Overwrites any existing file.
        pickle.dump(obj, outp, pickle.HIGHEST_PROTOCOL)
    # Text file to save the individuals
    text_file = filename + ".txt"
    with open(text_file, "a") as file:
        if obj.bestIsPerfect:
            file.write("\n\n\n Perfect solution " + filename + ": population_numDistinctPaths_m_n_k \n")
        else:
            file.write(f"\n\n\n Not a perfect solution, best fitness: {obj.best_fitness}\
                        {filename} :population_numDistinctPaths_m_n_k \n")
        paths = []   
        for alphabet_path in obj.individuals[obj.bfi].translate(obj.paths):
            paths.append("".join(alphabet_path))
        paths.sort()
        for index, path in enumerate(paths):
            file.write(path + "   " + str(obj.individuals[obj.bfi].sequences[index])+ '\n')
        file.write("\n\n\n")
    
# Creating/Opening an excel file to store the best configs when doing paralell
# search, again just in case shelve does not work

def read_object(filename, showBestIndividual=True, showTranslated=True):
    # Example file name == "population_5_4_2_3"
    with open(filename, "rb") as file:
        obj = pickle.load(file)
    if showBestIndividual and isinstance(obj, Population): # In case the object is a population
        obj.individuals[obj.bfi].show(obj.paths)
    if showTranslated:
        for list_path in obj.individuals[obj.bfi].translate(obj.paths):
            print("".join(list_path))
    return obj

def save_config_in_txt_file(config_integer, target, m, n, k,filename ="successful_config.txt"):
    with open(filename, "a") as file:
        file.write(f"{config_integer=} {target=} {m=} {n=} {k=} \n")
        
      
if __name__ == '__main__':
    # The core function of this data collection is parallel_search()
    # Note that parallel_search takes target,m,n, and k as arguments, hence
    # it is independent of the table we are trying to fill. It just needs the 
    # dimensions of the lattice and k

    # Hence, you can make this data collection way quicker if you ran the rows or individual entries(even better)
    # of the tables we are trying to generate by running parallel_search() in parallel with different arguments(the dimension
    # of the entry)
    for i in [2,3,4,5,6]:
        collect_data_greedy(i,i)
        collect_data_genetic(i,i)

















# Unused functions. Just during development

"""
def nbrr(non_bias_random_times):
    
    #Testing the run time of the evolutionary search using the 'non_bias_random' argument for kill_mode.
    #Returns a list of run times.
    
    for i in range(10):
        start_time = time.perf_counter()
        test(700, 15, 6, 3, 5, kill_mode="non_bias_random")
        end_time = time.perf_counter()
        non_bias_random_times.append((end_time - start_time) / 60)

    return non_bias_random_times


def kbr(kill_bottom_times):
    
    #Testing the run time of the evolutionary search using the 'kill_bottom' argument for kill_mode.
    #Returns a list of run times.
    
    for i in range(10):
        start_time = time.perf_counter()
        test(700, 15, 6, 3, 5, kill_mode="kill_bottom")
        end_time = time.perf_counter()
        kill_bottom_times.append((end_time - start_time) / 60)

    return kill_bottom_times


def compare_kill_mode():
    #Comparing runtime between 'non_random_bias' kill_mode and 'kill_bottom' kill mode.
    #No return value.
    
    kill_bottom_times = []
    non_bias_random_times = []
    kr = []
    nr = []
    with cf.ProcessPoolExecutor() as executor:
        f1 = executor.submit(kbr, kill_bottom_times)
        f2 = executor.submit(nbrr, non_bias_random_times)
        kr = f1.result()
        nr = f2.result()
    print("kr av: ", sum(kr) / len(kr))
    print("nr av: ", sum(nr) / len(nr))
"""
